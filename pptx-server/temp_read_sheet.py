import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\yuher\Projects\kizmeal-renewal\pptx-server\키즈밀_식단표_6월_테스트데이터_v5.4.xlsx', data_only=True)
print("시트 목록:", wb.sheetnames)

# '원 참고' 시트가 없으므로 '📋 원별 참조표' 시트 읽기
target = '📋 원별 참조표'
if target in wb.sheetnames:
    ws = wb[target]
    print(f"\n['{target}'] 행 수: {ws.max_row}, 열 수: {ws.max_column}\n")
    headers = [cell.value for cell in ws[1]]
    print("헤더:", headers)
    print("\n---데이터---")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            print(row)
else:
    print(f"'{target}' 시트도 없음")
