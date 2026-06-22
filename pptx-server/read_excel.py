# -*- coding: utf-8 -*-
"""
키즈밀 식단표 엑셀 파싱 + 49개원 PPTX 일괄 생성 (메인 러너)
========================================================================

실행:
    python read_excel.py

흐름:
  1) 엑셀 파싱 (5주 메뉴 + 49원 설정)
  2) 49개원 PPTX 생성 → output/{메인식단표원명}_202606.pptx
  3) PDF/JPG 변환 (LibreOffice 설치 시)
  4) ✅성공 / ⚠️스킵 / ❌실패 결과 출력

⚠️ 일본어 금지.
"""

import os
import re
import subprocess
import sys

import openpyxl
from bracket_parser import _BRACKET_RE, wants_fruit
from pptx_generator import generate

HERE = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH    = os.path.join(HERE, '키즈밀_식단표_6월_테스트데이터_v5.4.xlsx')
TEMPLATE_PATH = os.path.join(HERE, 'templates', '2026년 6월 식단표(양식).pptx')
OUTPUT_DIR    = os.path.join(HERE, 'output')

# ── 주차별 행 번호 (1-indexed) ─────────────────────────────────────
_ROW_HEADER     = 3
_ROW_BAP        = 4
_ROW_GUK        = 5
_ROW_BANCHAN1   = 6
_ROW_BANCHAN2   = 7
_ROW_BANCHAN3   = 8
_ROW_BANCHAN4   = 9
_ROW_BANCHAN5   = 10
_ROW_LUNCH_NUTRI= 11
_ROW_LUNCH_SPEC = 12   # 원별특이사항 (후식과일 + pm간식 오버라이드)
_ROW_LUNCH_EXC  = 13   # 예외규칙 (반찬 제거)
_ROW_AM_MENU    = 14
_ROW_AM_NUTRI   = 15
_ROW_AM_SPEC    = 16   # 오전 원별특이사항
_ROW_AM_EXC     = 17   # 오전 예외규칙
_ROW_BIRTHDAY   = 18
_ROW_PM_MENU    = 19
_ROW_PM_NUTRI   = 20
_ROW_PM_SPEC    = 21   # 오후 원별특이사항
_ROW_PM_EXC     = 22
_ROW_CARE_MENU  = 23
_ROW_CARE_NUTRI = 24

# ── 원별 참조표 컬럼 ──────────────────────────────────────────────
_COL_NO         = 1
_COL_NAME       = 2   # 원 명 (공식)
_COL_DISPLAY    = 3   # 식단표 표기명 (슬라이드 표시)
_COL_MAIN_NAME  = 4   # 메인식단표원명 (bracket 매칭용)
_COL_HAS_AM     = 5
_COL_HAS_PM     = 6
_COL_FILE_FMT   = 7   # PDF/JPG/PPT/PDF+JPG
_COL_ENG        = 8   # 영문 여부
_COL_NOTE       = 9
_COL_EMAIL      = 10
_COL_SLIDES     = 11  # 1P/2P/3P/4P
_COL_LABEL      = 12  # 오전/간식
_COL_CARE       = 13
_COL_BIRTHDAY   = 14
_COL_FRUIT      = 15  # O → add_fruit=True
_COL_ALLERGY    = 16

# ── 원별 참조표 설정 행 (원산지/원재료, B=라벨 C=값) ──────────────
_ROW_ORIGIN_BODY  = 64   # C64: 원산지 본문
_ROW_ORIGIN_DISC  = 65   # C65: 원산지 면책문구
_ROW_MATERIAL     = 66   # C66: 원재료 표시안내
_COL_SETTING_LABEL = 2   # B열 (라벨)
_COL_SETTING_VALUE = 3   # C열 (값)

# ── 동탄P 고정 오전간식 ────────────────────────────────────────────
_DONGTAN_FIXED_AM = {'menu': '유기농우유②', 'nutrition': ''}


# ════════════════════════════════════════════════════════════════════
# 1. 파싱 유틸
# ════════════════════════════════════════════════════════════════════
def _v(ws, row, col):
    """셀 값 문자열 변환 (None → '')."""
    val = ws.cell(row, col).value
    if val is None:
        return ''
    return str(val).strip()


def _is_skipped(header_str):
    """'해당없음' 또는 빈 헤더 = 해당 없는 날."""
    return not header_str or '해당없음' in header_str


def parse_header(s):
    """
    '월  6/1', '수  6/3\\n공휴일\\n(선거일)', '🎂 금  6/5' 등 헤더 파싱.
    반환: {'date': '01', 'is_holiday': bool, 'is_birthday': bool} or None
    """
    if _is_skipped(s):
        return None
    m = re.search(r'\d{1,2}/(\d{1,2})', s)
    if not m:
        return None
    return {
        'date': m.group(1).zfill(2),
        'is_holiday': '공휴일' in s or '휴원' in s,
        'is_birthday': '🎂' in s or '생일' in s,
    }


def _embed_brackets(base, *extras):
    """
    extra 텍스트의 모든 대괄호 표현을 base 뒤에 인라인 추가.
    [후식과일]은 제외 (별도 has_fruit 플래그로 처리).
    """
    base = '' if not base else str(base)
    parts = []
    for txt in extras:
        if not txt:
            continue
        for m in _BRACKET_RE.finditer(str(txt)):
            b = m.group(0)
            inner = m.group(1).strip()
            if inner != '후식과일':
                parts.append(b)
    return base + ''.join(parts)


def _inject_exception_to_banchan(banchans, exc_text):
    """
    예외규칙 텍스트에서 [원명-메뉴X] 추출 후 해당 반찬에 인라인 삽입.
    예: '[잉파-다시마튀각X]' → banchan3='다시마튀각' 에 '...[잉파-다시마튀각X]' 추가
    """
    if not exc_text:
        return
    keys = ['banchan1', 'banchan2', 'banchan3', 'banchan4', 'banchan5']
    for m in _BRACKET_RE.finditer(str(exc_text)):
        content = m.group(1)
        bracket = m.group(0)
        if '-' not in content:
            continue
        _, menu_part = content.split('-', 1)
        if not menu_part.endswith('X'):
            continue
        target = menu_part[:-1].strip()
        for key in keys:
            val = banchans.get(key, '')
            if val and target in str(val):
                banchans[key] = str(val) + bracket
                break


# ════════════════════════════════════════════════════════════════════
# 2. 주차 파싱
# ════════════════════════════════════════════════════════════════════
def parse_week(ws, wk_num):
    """
    주차 시트 파싱 → {'wk': int, 'days': [day_dict*5]}
    day_dict: {'date','is_holiday','is_birthday','lunch',
               'morning_snack','afternoon_snack','care_snack'}
    is_skipped=True 이면 나머지 키 없음.
    """
    days = []
    for col in range(2, 7):  # 월~금
        hdr_raw = ws.cell(_ROW_HEADER, col).value
        hdr_str = '' if hdr_raw is None else str(hdr_raw).strip()
        hdr = parse_header(hdr_str)
        if hdr is None:
            days.append({'is_skipped': True})
            continue

        # ── 중식 데이터 ────────────────────────────────────────────
        bap = _v(ws, _ROW_BAP, col)
        guk = _v(ws, _ROW_GUK, col)
        banchans = {
            'banchan1': _v(ws, _ROW_BANCHAN1, col),
            'banchan2': _v(ws, _ROW_BANCHAN2, col),
            'banchan3': _v(ws, _ROW_BANCHAN3, col),
            'banchan4': _v(ws, _ROW_BANCHAN4, col),
            'banchan5': _v(ws, _ROW_BANCHAN5, col),
        }
        lunch_nutri = _v(ws, _ROW_LUNCH_NUTRI, col)
        lunch_spec  = _v(ws, _ROW_LUNCH_SPEC, col)
        lunch_exc   = _v(ws, _ROW_LUNCH_EXC, col)

        # 예외규칙 반찬에 인라인 삽입
        if lunch_exc:
            _inject_exception_to_banchan(banchans, lunch_exc)

        # bap/guk에 lunch_spec 중 기타 오버라이드 삽입 (반찬 이외 교체는 없음)
        # → 중식 원별특이사항의 비-후식과일 bracket은 pm_snack으로 처리
        lunch = {
            'bap':       bap,
            'guk':       guk,
            'banchan1':  banchans['banchan1'],
            'banchan2':  banchans['banchan2'],
            'banchan3':  banchans['banchan3'],
            'banchan4':  banchans['banchan4'],
            'banchan5':  banchans['banchan5'],
            'nutrition': lunch_nutri,
            'has_fruit': wants_fruit(lunch_spec),
        }

        # ── 오전간식 ───────────────────────────────────────────────
        am_menu_raw = _v(ws, _ROW_AM_MENU, col)
        am_spec     = _v(ws, _ROW_AM_SPEC, col)
        am_exc      = _v(ws, _ROW_AM_EXC, col)
        am_menu     = _embed_brackets(am_menu_raw, am_spec, am_exc)
        morning_snack = {
            'menu':      am_menu,
            'nutrition': _v(ws, _ROW_AM_NUTRI, col),
        }

        # ── 오후간식 (중식 원별특이사항 비-과일 bracket 추가 포함) ──
        pm_menu_raw = _v(ws, _ROW_PM_MENU, col)
        pm_spec     = _v(ws, _ROW_PM_SPEC, col)
        pm_exc      = _v(ws, _ROW_PM_EXC, col)
        pm_menu     = _embed_brackets(pm_menu_raw, pm_spec, pm_exc, lunch_spec)
        afternoon_snack = {
            'menu':      pm_menu,
            'nutrition': _v(ws, _ROW_PM_NUTRI, col),
        }

        # ── 돌봄간식 ───────────────────────────────────────────────
        care_snack = {
            'menu':      _v(ws, _ROW_CARE_MENU, col),
            'nutrition': _v(ws, _ROW_CARE_NUTRI, col),
        }

        days.append({
            'date':           hdr['date'],
            'is_holiday':     hdr['is_holiday'] or not bool(bap),
            'is_birthday':    hdr['is_birthday'],
            'is_skipped':     False,
            'lunch':          lunch,
            'morning_snack':  morning_snack,
            'afternoon_snack':afternoon_snack,
            'care_snack':     care_snack,
        })

    return {'wk': wk_num, 'days': days}


# ════════════════════════════════════════════════════════════════════
# 3. 원별 설정 파싱
# ════════════════════════════════════════════════════════════════════
def determine_type(slides_str, has_am, has_pm, care, note, eng):
    """타입코드 결정."""
    n = int(str(slides_str or '1')[0])

    if care == 'O':
        return 'songpaE'

    if n == 4:
        return 'F' if eng == 'O' else 'G'

    if n == 3:
        return 'E'

    if n == 2:
        if note and 'Yonder' in str(note):
            return 'D'
        return 'B'

    # 1P
    if has_am == 'O' and has_pm == 'O':
        return 'C'

    return 'A'


def parse_branches(ws):
    """원별 참조표 파싱 → 49개 cfg 딕트 리스트."""
    branches = []
    for row in range(4, 60):
        no = ws.cell(row, _COL_NO).value
        if no is None:
            break

        name        = ws.cell(row, _COL_MAIN_NAME).value or ''
        display     = ws.cell(row, _COL_DISPLAY).value or name
        email       = ws.cell(row, _COL_EMAIL).value or ''
        slides_str  = ws.cell(row, _COL_SLIDES).value or '1P'
        has_am      = ws.cell(row, _COL_HAS_AM).value
        has_pm      = ws.cell(row, _COL_HAS_PM).value
        care        = ws.cell(row, _COL_CARE).value
        label       = ws.cell(row, _COL_LABEL).value or '오전'
        fruit       = ws.cell(row, _COL_FRUIT).value
        note        = ws.cell(row, _COL_NOTE).value
        eng         = ws.cell(row, _COL_ENG).value
        file_fmt    = ws.cell(row, _COL_FILE_FMT).value or 'PDF'

        type_code = determine_type(slides_str, has_am, has_pm, care, note, eng)

        cfg = {
            'name':         str(name).strip(),
            'display_name': str(display).strip(),
            'email':        str(email).strip(),
            'type':         type_code,
            'snack_label':  str(label).strip(),
            'add_fruit':    (str(fruit).strip() == 'O'),
            'fruit_text':   '제철과일',
            'file_fmt':     str(file_fmt).strip(),
            'use_pm_as_am': (has_am != 'O' and has_pm == 'O'),
            'fixed_am':     None,
        }

        # 동탄P: 오전간식 항상 유기농우유② 고정
        if name == '동탄P':
            cfg['fixed_am'] = _DONGTAN_FIXED_AM

        branches.append(cfg)

    return branches


# ════════════════════════════════════════════════════════════════════
# 4. 엑셀 전체 로드
# ════════════════════════════════════════════════════════════════════
def load_excel(path):
    """
    Returns:
      menu_data: {'year':2026, 'month':6, 'weeks':{'1':..., '5':...}}
      branches:  list of cfg dicts
      date_map:  {'09':'09', '10':'10', '16':'16', '17':'17'} (6월 고정)
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # 주차별 메뉴 파싱
    wk_names = ['1주차', '2주차', '3주차', '4주차', '5주차']
    weeks = {}
    for i, sname in enumerate(wk_names, start=1):
        ws = wb[sname]
        wk_data = parse_week(ws, i)
        weeks[str(i)] = wk_data

    menu_data = {'year': 2026, 'month': 6, 'weeks': weeks}

    # 날짜 그룹 대체용 date_map (2주화/수, 3주화/수)
    def _day_date(wk_str, day_idx):
        days = weeks.get(wk_str, {}).get('days', [])
        if day_idx < len(days) and days[day_idx] and not days[day_idx].get('is_skipped'):
            return days[day_idx].get('date', '')
        return ''

    date_map = {
        '09': _day_date('2', 1),   # 2주 화
        '10': _day_date('2', 2),   # 2주 수
        '16': _day_date('3', 1),   # 3주 화
        '17': _day_date('3', 2),   # 3주 수
    }

    # 원별 참조표 파싱
    ws_branches = wb['📋 원별 참조표']
    branches = parse_branches(ws_branches)

    # 원산지·원재료 텍스트 읽기 (C64~C66, 없거나 비어있으면 None)
    try:
        def _setting(row):
            val = ws_branches.cell(row, _COL_SETTING_VALUE).value
            return str(val).strip() if val else None

        origin_body = _setting(_ROW_ORIGIN_BODY)
        if origin_body:
            menu_data['origin_text'] = {
                'body':       origin_body,
                'disclaimer': _setting(_ROW_ORIGIN_DISC),
            }
        else:
            menu_data['origin_text'] = None

        menu_data['material_text'] = _setting(_ROW_MATERIAL)
    except Exception:
        menu_data['origin_text']   = None
        menu_data['material_text'] = None

    return menu_data, branches, date_map


# ════════════════════════════════════════════════════════════════════
# 5. LibreOffice 변환
# ════════════════════════════════════════════════════════════════════
def convert_pptx(pptx_path, fmt):
    """
    LibreOffice headless 변환.
    fmt: 'PDF', 'JPG', 'PDF+JPG', 'PPT' (PPT는 변환 없음)
    Returns: list of 생성된 파일 경로
    """
    if 'PPT' == fmt.upper():
        return [pptx_path]

    formats = []
    if 'PDF' in fmt.upper():
        formats.append(('pdf', 'impress_pdf_Export'))
    if 'JPG' in fmt.upper():
        formats.append(('jpg', 'impress_jpg_Export'))

    results = []
    out_dir = os.path.dirname(pptx_path)

    for ext, filter_name in formats:
        try:
            subprocess.run(
                [
                    'soffice', '--headless',
                    f'--convert-to', f'{ext}:{filter_name}',
                    '--outdir', out_dir,
                    pptx_path,
                ],
                check=True,
                capture_output=True,
                timeout=120,
            )
            base = os.path.splitext(pptx_path)[0]
            out_file = f'{base}.{ext}'
            if os.path.exists(out_file):
                results.append(out_file)
        except FileNotFoundError:
            print('  ⚠️  LibreOffice 미설치 → 변환 건너뜀')
            return None
        except subprocess.CalledProcessError as e:
            print(f'  ⚠️  LibreOffice 변환 실패: {e.stderr.decode("utf-8", errors="replace")}')
        except subprocess.TimeoutExpired:
            print('  ⚠️  LibreOffice 변환 타임아웃')

    return results


# ════════════════════════════════════════════════════════════════════
# 6. 메인 생성 루프
# ════════════════════════════════════════════════════════════════════
def main():
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    print('=' * 65)
    print('키즈밀 식단표 PPTX 일괄 생성 (2026년 6월)')
    print('=' * 65)

    # 엑셀 파싱
    if not os.path.exists(EXCEL_PATH):
        print(f'❌ 엑셀 파일 없음: {EXCEL_PATH}')
        return 1
    if not os.path.exists(TEMPLATE_PATH):
        print(f'❌ 빈 양식 없음: {TEMPLATE_PATH}')
        return 1

    print(f'\n[1/3] 엑셀 파싱 중...')
    menu_data, branches, date_map = load_excel(EXCEL_PATH)
    print(f'  ✅ {len(branches)}개 원 설정 로드')
    print(f'  날짜맵: {date_map}')

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 생성 루프
    print(f'\n[2/3] PPTX 생성 중...')
    n_ok = n_skip = n_fail = 0
    results = []

    for cfg in branches:
        name = cfg['name']
        out_pptx = os.path.join(OUTPUT_DIR, f'{name}_202606.pptx')

        try:
            generate(cfg, menu_data, TEMPLATE_PATH, out_pptx, date_map=date_map)
            n_ok += 1
            results.append(('OK', name, out_pptx, cfg))
            print(f'  ✅ OK      [{cfg["type"]:8s}] {name}')
        except Exception as exc:
            n_fail += 1
            results.append(('FAIL', name, None, cfg))
            print(f'  ❌ FAIL    [{cfg["type"]:8s}] {name} — {type(exc).__name__}: {exc}')

    # 변환
    print(f'\n[3/3] PDF/JPG 변환 중...')
    lo_missing = False
    for status, name, pptx_path, cfg in results:
        if status != 'OK' or not pptx_path:
            continue
        fmt = cfg.get('file_fmt', 'PDF')
        if fmt.upper() == 'PPT':
            print(f'  ⏭️  SKIP   {name} (PPT 형식, 변환 불필요)')
            continue
        if lo_missing:
            continue
        converted = convert_pptx(pptx_path, fmt)
        if converted is None:
            lo_missing = True
        elif converted:
            fnames = ', '.join(os.path.basename(p) for p in converted)
            print(f'  ✅ OK      {name} → {fnames}')

    # 요약
    print('\n' + '=' * 65)
    print(f'생성 결과: ✅성공 {n_ok}  /  ⚠️스킵 {n_skip}  /  ❌실패 {n_fail}  (전체 {len(branches)}개)')
    if n_fail > 0:
        print('실패 원 목록:')
        for status, name, _, _ in results:
            if status == 'FAIL':
                print(f'  - {name}')
    print('=' * 65)
    return 0 if n_fail == 0 else 1


if __name__ == '__main__':
    sys.exit(main())
