# -*- coding: utf-8 -*-
"""커밋2 — 기준폼 v6 첫인상 개선 (일회성 스크립트).

배서영 영양사가 v6를 처음 열었을 때의 가독성을 고친다. 값(메뉴·안내문)은
건드리지 않고 서식만 바꾼다. 실행 후 삭제해도 되는 일회성 도구다.

■ 왜 조건 기반으로 고르는가
좌표를 하드코딩하면 주차마다 다른 상태(공휴일 표본·해당없음 표본)까지
싸잡아 덮어쓴다. 그래서 "지금 어떤 서식인가"를 보고 대상만 고른다.

■ 손대지 않는 것
  · 열 폭 — B~F가 이미 28.0으로 통일돼 있다(HANDOFF 전제 정정)
  · 18행 생일간식 — 8pt/'—'(비우는 칸)와 9pt/212121(입력 칸)의
    의도된 2상태 체계다
  · 1주차 D열·5주차 D열 — gen_form.py 의 공휴일·해당없음 서식 표본
  · 시트 숨김 — 커밋3(가이드 두 층 분리)과 한 몸이라 그때 함께
"""
import copy
import openpyxl
from openpyxl.styles import Font

BASE = '키즈밀_식단표_기준폼_v6.xlsx'
WEEK_SHEETS = ['1주차', '2주차', '3주차', '4주차', '5주차']

# 반찬5 입력 칸의 "고쳐야 할 상태" — 커밋① 이후 남은 잔재
SIDE5_OLD_SZ, SIDE5_OLD_COLOR = 7.0, 'FFBDBDBD'
# 정상 입력 칸 기준값 (반찬1~4 실측)
BODY_SZ, BODY_COLOR = 9.0, 'FF212121'

ROW_SIDE5 = 10
ROW_LABELS = range(4, 25)          # A4~A24 (A3 '구분' 10pt 은 제외)
ROWS_FREE_HEIGHT = (12, 16, 21)    # 원별특이사항
ROW_NOTE = 2                       # D2 빨간 안내문
NOTE_HEIGHT = 36.75                # 1주차 기준. 2~5주차를 여기에 맞춘다
NOTE_SZ = 10.0
LEGEND_ROWS = (25, 26)
LEGEND_SZ, LEGEND_COLOR = 9.0, 'FF424242'
LABEL_SZ = 9.0
DEFAULT_ROW_HEIGHT = 24.0
DROP_FROM_A25 = '  ■ 반찬5=없으면 빈칸'


def resize(cell, sz=None, color=None):
    """폰트 크기·색만 바꾸고 나머지 속성(굵기·서체)은 보존한다.

    ⚠️ cell.font 는 StyleProxy 라 속성을 직접 대입할 수 없다. copy 후 교체.
    """
    f = copy.copy(cell.font)
    cell.font = Font(
        name=f.name,
        sz=sz if sz is not None else f.sz,
        b=f.b, i=f.i, u=f.u, strike=f.strike,
        color=color if color is not None else f.color,
        vertAlign=f.vertAlign,
    )


def main():
    wb = openpyxl.load_workbook(BASE)
    n = dict.fromkeys(
        ['반찬5', 'A열라벨', '2행높이', 'D2', '범례', '높이해제', '기본높이', '인쇄'], 0)

    for name in WEEK_SHEETS:
        ws = wb[name]

        # ① 반찬5 입력 칸 — 커밋①이 문구만 지우고 남긴 7pt·연회색을 되돌린다
        for col in range(2, 7):
            # 값 유무는 보지 않는다. 값으로 거르면 이미 작성 중인 파일에서
            # 입력된 칸만 쏙 빠진다. 공휴일·해당없음 표본(9pt/9E9E9E)은
            # 색·크기 조건에서 자연히 걸러진다.
            c = ws.cell(ROW_SIDE5, col)
            same_color = c.font.color and c.font.color.rgb == SIDE5_OLD_COLOR
            if c.font.sz == SIDE5_OLD_SZ and same_color:
                resize(c, sz=BODY_SZ, color=BODY_COLOR)
                n['반찬5'] += 1

        # ② A열 행 라벨 8pt → 9pt (색은 행마다 달라 보존)
        for row in ROW_LABELS:
            c = ws.cell(row, 1)
            if c.value is not None and c.font.sz == 8.0:
                resize(c, sz=LABEL_SZ)
                n['A열라벨'] += 1

        # ③ 2행 높이 통일 — 2~5주차가 18.0 이라 안내문이 지금도 잘려 있다
        if ws.row_dimensions[ROW_NOTE].height != NOTE_HEIGHT:
            ws.row_dimensions[ROW_NOTE].height = NOTE_HEIGHT
            n['2행높이'] += 1

        # ④ D2 안내문 8pt → 10pt (D2:F2 병합 → 좌상단만 유효)
        if ws['D2'].font.sz != NOTE_SZ:
            resize(ws['D2'], sz=NOTE_SZ)
            n['D2'] += 1

        # ⑤ A25 중복 문구 삭제 — 같은 내용이 A26 앞머리에 또 있다
        if ws['A25'].value and DROP_FROM_A25 in ws['A25'].value:
            ws['A25'] = ws['A25'].value.replace(DROP_FROM_A25, '')
            n['범례'] += 1

        # ⑥ 범례 7pt/616161 → 9pt/424242
        for row in LEGEND_ROWS:
            resize(ws.cell(row, 1), sz=LEGEND_SZ, color=LEGEND_COLOR)

        # ⑦ 원별특이사항 행 높이 고정 해제 → 내용 길이에 따라 자동
        for row in ROWS_FREE_HEIGHT:
            # customHeight 는 읽기 전용이다. height=None 이면 openpyxl 이
            # ht/customHeight 속성을 아예 쓰지 않아 자동 높이가 된다.
            ws.row_dimensions[row].height = None
            n['높이해제'] += 1

        # ⑧ 빈 폼이 납작해지지 않도록 자동 높이의 출발점을 올린다
        ws.sheet_format.defaultRowHeight = DEFAULT_ROW_HEIGHT
        ws.sheet_format.customHeight = True
        n['기본높이'] += 1

        # ⑨ 인쇄: 가로 1장 유지, 세로 제한 해제 → 세로 축소가 사라진다
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        n['인쇄'] += 1

    wb.save(BASE)
    for k, v in n.items():
        print(f'  {k}: {v}개소')


if __name__ == '__main__':
    main()
